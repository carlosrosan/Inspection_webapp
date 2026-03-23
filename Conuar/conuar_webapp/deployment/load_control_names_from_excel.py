#!/usr/bin/env python3
"""
Reads "Info de fotos.xlsx" and outputs SQL INSERT statements for the control_names table.
Columns: control_id (ID), control_name (Medición), valor_esperado, tolerancia,
separador, barra, pos_z_carro, pos_x_plato. Multiple rows per control_id are allowed.

Usage:
  python load_control_names_from_excel.py                    # print INSERTs to stdout
  python load_control_names_from_excel.py --output out.sql   # write to file
  python load_control_names_from_excel.py --django            # insert via Django DB
  python load_control_names_from_excel.py --list-columns      # print Excel column names
"""

import argparse
import os
import re
import sys

EXCEL_FILENAME = "Info de fotos.xlsx"
ID_COL = "ID"
MEDICION_COL = "Medicion"  # Excel may show "Medición"

# All columns in control_names (must match control_names.sql, without id)
DB_COLUMNS = [
    "control_id",
    "control_name",
    "valor_esperado",
    "tolerancia",
    "separador",
    "barra",
    "pos_z_carro",
    "pos_x_plato",
]

# Possible Excel header names for each column (case-insensitive match)
EXCEL_HEADER_MAP = {
    "control_id": ["id"],
    "control_name": ["medicion", "medición"],
    "valor_esperado": ["valor esperado", "valor_esperado"],
    "tolerancia": ["tolerancia"],
    "separador": ["separador"],
    "barra": ["barra"],
    "pos_z_carro": ["pos z carro", "pos_z_carro", "pos z"],
    "pos_x_plato": ["pos x plato", "pos_x_plato", "pos x"],
}


def _normalize_header(h):
    if h is None:
        return ""
    s = str(h).strip()
    return re.sub(r"\s+", " ", s).lower().replace("í", "i").replace("ó", "o")


def _find_column_index(header_list, possible_names):
    """Return index of first header that matches any of possible_names (normalized)."""
    normalized = [_normalize_header(h) for h in header_list]
    for name in possible_names:
        n = _normalize_header(name)
        for i, h in enumerate(normalized):
            if n in h or h == n:
                return i
    return None


def _excel_headers_to_columns(header_list):
    """Build mapping: db_column -> excel column index (0-based)."""
    header = [str(h).strip() if h is not None else "" for h in header_list]
    try:
        idx_id = header.index(ID_COL)
    except ValueError:
        idx_id = _find_column_index(header_list, ["id"])
        if idx_id is None:
            idx_id = 0
    try:
        idx_med = header.index(MEDICION_COL)
    except ValueError:
        try:
            idx_med = header.index("Medición")
        except ValueError:
            idx_med = _find_column_index(header_list, ["medicion", "medición"])
            idx_med = idx_med if idx_med is not None else 1

    col_map = {"control_id": idx_id, "control_name": idx_med}
    for db_col in DB_COLUMNS:
        if db_col in col_map:
            continue
        idx = _find_column_index(header_list, EXCEL_HEADER_MAP.get(db_col, [db_col.replace("_", " ")]))
        if idx is not None:
            col_map[db_col] = idx
    return col_map


def _cell_str(val):
    if val is None:
        return ""
    return str(val).strip()


def read_excel(excel_path):
    """
    Return (list of row dicts, ordered column names for DB).
    Each row dict has keys from DB_COLUMNS (only keys present in col_map).
    """
    try:
        import openpyxl
    except ImportError:
        try:
            import pandas as pd
            df = pd.read_excel(excel_path)
            df.columns = [str(c).strip() for c in df.columns]
            col_map = _excel_headers_to_columns(list(df.columns))
            if "control_id" not in col_map or "control_name" not in col_map:
                print("Could not find ID and Medicion columns. Found:", list(df.columns), file=sys.stderr)
                return [], []
            db_columns = [c for c in DB_COLUMNS if c in col_map]
            rows = []
            for _, r in df.iterrows():
                id_val = r.iloc[col_map["control_id"]] if col_map["control_id"] < len(r) else None
                if id_val is None or _cell_str(id_val) == "":
                    continue
                rec = {}
                for col in db_columns:
                    idx = col_map[col]
                    rec[col] = _cell_str(r.iloc[idx]) if idx < len(r) else ""
                rows.append(rec)
            return rows, db_columns
        except ImportError:
            print("Install openpyxl or pandas: pip install openpyxl pandas", file=sys.stderr)
            return [], []

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        wb.close()
        return [], []
    col_map = _excel_headers_to_columns(header)
    db_columns = [c for c in DB_COLUMNS if c in col_map]
    if "control_id" not in col_map or "control_name" not in col_map:
        wb.close()
        print("Could not find ID and Medicion columns. Found:", list(header), file=sys.stderr)
        return [], []
    rows = []
    for row in rows_iter:
        if not row:
            continue
        id_val = row[col_map["control_id"]] if col_map["control_id"] < len(row) else None
        if id_val is None or _cell_str(id_val) == "":
            continue
        rec = {}
        for col in db_columns:
            idx = col_map[col]
            rec[col] = _cell_str(row[idx]) if idx < len(row) else ""
        rows.append(rec)
    wb.close()
    return rows, db_columns


def escape_sql(s):
    return s.replace("'", "''")


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    default_path = os.path.join(script_dir, EXCEL_FILENAME)
    parser = argparse.ArgumentParser(description="Load control_names from Excel into SQL or Django DB.")
    parser.add_argument("--excel", default=default_path, help=f"Path to Excel file (default: {EXCEL_FILENAME})")
    parser.add_argument("--output", "-o", help="Write INSERTs to this file instead of stdout")
    parser.add_argument("--django", action="store_true", help="Insert into DB using Django (run from project root)")
    parser.add_argument("--list-columns", action="store_true", help="Print Excel column names and exit")
    args = parser.parse_args()

    if not os.path.isfile(args.excel):
        print(f"File not found: {args.excel}", file=sys.stderr)
        return 1

    data, db_columns = read_excel(args.excel)
    if args.list_columns:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(args.excel, read_only=True)
            header = next(wb.active.iter_rows(values_only=True), None)
            wb.close()
            print("Excel columns:", list(header) if header else [])
        except Exception as e:
            print(e, file=sys.stderr)
        return 0

    if not data:
        print("No rows found in Excel.", file=sys.stderr)
        return 1

    if args.django:
        try:
            parent = os.path.dirname(script_dir)
            if parent not in sys.path:
                sys.path.insert(0, parent)
            os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
            import django
            django.setup()
            from django.db import connection
            vendor = connection.vendor
            placeholders_sqlite = ", ".join("?" * len(db_columns))
            placeholders_pg = ", ".join(["%s"] * len(db_columns))
            cols = ", ".join(db_columns)
            with connection.cursor() as c:
                c.execute("DELETE FROM control_names")
                for row in data:
                    values = [row.get(col, "") for col in db_columns]
                    if vendor == "sqlite3":
                        c.execute(
                            f"INSERT INTO control_names ({cols}) VALUES ({placeholders_sqlite})",
                            values,
                        )
                    else:
                        c.execute(
                            f"INSERT INTO control_names ({cols}) VALUES ({placeholders_pg})",
                            values,
                        )
            print(f"Inserted {len(data)} rows into control_names.", file=sys.stderr)
        except Exception as e:
            print(f"Django insert failed: {e}", file=sys.stderr)
            return 1
        return 0

    lines = ["DELETE FROM control_names;", ""]
    for row in data:
        vals = ", ".join(f"'{escape_sql(row.get(col, ''))}'" for col in db_columns)
        cols = ", ".join(db_columns)
        lines.append(f"INSERT INTO control_names ({cols}) VALUES ({vals});")
    out = "\n".join(lines)
    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(out)
        print(f"Wrote {len(data)} INSERTs to {args.output}", file=sys.stderr)
    else:
        print(out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
